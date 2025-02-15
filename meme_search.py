import requests
import json
from googleapiclient.discovery import build
import pandas as pd
from datetime import datetime, timedelta
import streamlit as st
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpxImage
import re
import time
import os
import zipfile
import yt_dlp
import random

# API 키를 st.secrets에서 가져오기
SERPAPI_API_KEY = st.secrets["SERPAPI_API_KEY"]
YOUTUBE_API_KEYS = st.secrets["YOUTUBE_API_KEYS"]

def get_youtube_api_key():
    """YouTube API 키를 순환하여 사용"""
    return random.choice(YOUTUBE_API_KEYS)

# 국가별 검색어 설정
COUNTRY_QUERIES = {
    "한국": {"term": "밈", "geo": "KR", "hl": "ko"},
    "미국": {"term": "meme", "geo": "US", "hl": "en"},
    "일본": {"term": "ミーム", "geo": "JP", "hl": "ja"}
}

# 기간 설정
PERIOD_OPTIONS = {
    "최근 1일": 1,
    "최근 1주일": 7,
    "최근 1달": 30,
    "최근 1년": 365
}

def parse_iso8601_duration(duration_str):
    """YouTube 동영상 길이를 초 단위로 변환"""
    pattern = r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?'
    match = re.match(pattern, duration_str)
    hours = int(match.group(1)) if match and match.group(1) else 0
    minutes = int(match.group(2)) if match and match.group(2) else 0
    seconds = int(match.group(3)) if match and match.group(3) else 0
    return hours * 3600 + minutes * 60 + seconds

def get_youtube_shorts(search_query, days_ago=365, max_results=5):
    """YouTube Shorts 검색 함수"""
    try:
        youtube = build('youtube', 'v3', developerKey=get_youtube_api_key())
        
        # 선택된 기간 전 날짜 계산
        period_ago = (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%dT%H:%M:%SZ')
        
        # 검색 실행
        search_response = youtube.search().list(
            q=f"{search_query} #shorts",
            part="id,snippet",
            maxResults=50,  # 더 많은 결과를 가져와서 필터링
            type="video",
            videoDuration="short",
            order="viewCount",
            publishedAfter=period_ago
        ).execute()

        videos = []
        video_ids = [item['id']['videoId'] for item in search_response['items']]

        # 비디오 상세 정보 가져오기
        video_response = youtube.videos().list(
            part="snippet,statistics,contentDetails",
            id=','.join(video_ids)
        ).execute()

        for video in video_response['items']:
            duration = parse_iso8601_duration(video['contentDetails']['duration'])
            
            # 60초 이하 동영상만 필터링
            if duration <= 60:
                video_info = {
                    'title': video['snippet']['title'],
                    'video_id': video['id'],
                    'view_count': int(video['statistics'].get('viewCount', 0)),
                    'like_count': int(video['statistics'].get('likeCount', 0)),
                    'published_at': video['snippet']['publishedAt'],
                    'channel_title': video['snippet']['channelTitle'],
                    'url': f'https://www.youtube.com/watch?v={video["id"]}'
                }
                videos.append(video_info)

        # 조회수 기준 상위 5개 반환
        return sorted(videos, key=lambda x: x['view_count'], reverse=True)[:max_results]

    except Exception as e:
        print(f"YouTube API 오류: {str(e)}")
        return []

def main():
    # 전체 화면 모드 설정
    st.set_page_config(layout="wide")
    
    # 타이틀 표시
    st.title("밈(Meme) 트렌드 및 YouTube Shorts 분석", anchor=False)
    
    # 다시 시작하기 버튼을 타이틀 아래에 왼쪽 정렬로 배치
    col1, col2, col3 = st.columns([1, 4, 4])
    with col1:
        st.button("🔄 다시 시작", on_click=lambda: [st.session_state.clear(), st.experimental_rerun()])
    
    # 세션 상태 초기화
    if 'results' not in st.session_state:
        st.session_state.results = {}
    if 'youtube_results' not in st.session_state:
        st.session_state.youtube_results = {}
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    
    # 국가별 국기 이모지 매핑
    COUNTRY_FLAGS = {
        "한국": "🇰🇷",
        "미국": "🇺🇸",
        "일본": "🇯🇵"
    }
    
    with st.expander("프로그램 설명", expanded=True):
        st.markdown("""
        이 프로그램은 다음 기능을 제공합니다:
        1. 선택한 국가의 밈 관련 급상승 밈 키워드를 수집
        2. 해당 밈 키워드로 YouTube Shorts 영상을 검색
        3. 결과를 엑셀 파일로 다운로드
        """)

    # 사이드바에 설정 옵션 추가
    st.sidebar.header("검색 설정")
    selected_countries = st.sidebar.multiselect(
        "국가 선택",
        options=list(COUNTRY_QUERIES.keys()),
        default=["한국"]
    )
    
    selected_period = st.sidebar.selectbox(
        "검색 기간",
        options=list(PERIOD_OPTIONS.keys()),
        index=3  # 기본값: 1년
    )

    if st.sidebar.button("분석 시작"):
        results = {}
        youtube_results = {}
        
        # 진행 상황을 보여줄 progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for idx, country in enumerate(selected_countries):
            try:
                q = COUNTRY_QUERIES[country]
                status_text.text(f"검색 중: {q['term']} ({q['geo']})")
                
                # Google Trends API 호출
                params = {
                    "engine": "google_trends",
                    "q": q["term"],
                    "geo": q["geo"],
                    "hl": q["hl"],
                    "data_type": "RELATED_QUERIES",
                    "api_key": SERPAPI_API_KEY
                }
                
                response = requests.get("https://serpapi.com/search.json", params=params)
                response.raise_for_status()
                data = response.json()
                
                related_queries = data.get("related_queries", {})
                rising_queries = related_queries.get("rising", [])[:15] if related_queries.get("rising") else []
                
                results[country] = {
                    "rising_related_queries": [{
                        "query": item["query"],
                        "value": item["value"]
                    } for item in rising_queries]
                }
                
                # YouTube Shorts 검색
                status_text.text(f"YouTube Shorts 검색 중... ({country})")
                youtube_results[country] = {}
                for item in rising_queries:
                    shorts_data = get_youtube_shorts(
                        item["query"],
                        days_ago=PERIOD_OPTIONS[selected_period]
                    )
                    youtube_results[country][item["query"]] = shorts_data
                
            except Exception as e:
                st.error(f"오류 발생 ({country}): {str(e)}")
                results[country] = {"error": str(e)}
            
            # 진행률 업데이트
            progress_bar.progress((idx + 1) / len(selected_countries))
        
        # 세션 상태 업데이트
        st.session_state.results = results
        st.session_state.youtube_results = youtube_results
        st.session_state.analysis_complete = True
        
        status_text.text("분석 완료!")
        
        # 엑셀 다운로드 버튼을 분석 완료 메시지 바로 아래에 배치
        try:
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # YouTube Shorts 결과 시트 (먼저 작성)
                youtube_data = []
                for country, queries in st.session_state.youtube_results.items():
                    for query, shorts in queries.items():
                        for short in shorts:
                            # 썸네일 URL 가져오기
                            try:
                                thumbnail_url = f"https://i.ytimg.com/vi/{short['video_id']}/hqdefault.jpg"
                            except:
                                thumbnail_url = ""

                            youtube_data.append({
                                "Country": country,
                                "Search Query": query,
                                "Title": short["title"],
                                "Channel": short["channel_title"],
                                "Views": short["view_count"],
                                "Likes": short["like_count"],
                                "Published Date": datetime.strptime(short["published_at"], "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d"),
                                "url": short["url"],
                                "thumbnail_url": thumbnail_url
                            })
                
                df_youtube = pd.DataFrame(youtube_data)
                df_youtube.to_excel(writer, sheet_name='YouTube Shorts', index=False)

                # Google Trends 결과 시트 (나중에 작성)
                trends_data = []
                for country, data in st.session_state.results.items():
                    for query in data.get("rising_related_queries", []):
                        trends_data.append({
                            "Country": country,
                            "Related Query": query["query"],
                            "Value": query["value"]
                        })
                
                df_trends = pd.DataFrame(trends_data)
                df_trends.to_excel(writer, sheet_name='Meme Keyword', index=False)

                # 스타일 적용
                workbook = writer.book
                for sheet_name in ['YouTube Shorts', 'Meme Keyword']:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 스타일
                    header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    header_font = Font(bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 전체 셀 중앙 정렬
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    # 열 너비 조정
                    if sheet_name == 'YouTube Shorts':
                        # 기본 열 너비 설정
                        for col_letter in ["A", "B"]:  # Country, Search Query
                            worksheet.column_dimensions[col_letter].width = 15
                        worksheet.column_dimensions["C"].width = 30  # Title
                        worksheet.column_dimensions["D"].width = 20  # Channel
                        for col_letter in ["E", "F"]:  # Views, Likes
                            worksheet.column_dimensions[col_letter].width = 12
                        worksheet.column_dimensions["G"].width = 15  # Published Date
                        worksheet.column_dimensions["H"].width = 7   # URL
                        worksheet.column_dimensions["I"].width = 30  # Thumbnail

                        # 행 높이 조정
                        worksheet.row_dimensions[1].height = 30
                        for row_idx in range(2, len(df_youtube) + 2):
                            worksheet.row_dimensions[row_idx].height = 135

                        # 썸네일 이미지 삽입
                        for i, row in df_youtube.iterrows():
                            try:
                                cell_row = i + 2  # 2행부터 시작
                                thumb_url = row["thumbnail_url"]
                                if thumb_url and isinstance(thumb_url, str):
                                    resp = requests.get(thumb_url, timeout=5)
                                    if resp.status_code == 200:
                                        img_data = BytesIO(resp.content)
                                        img = OpxImage(img_data)
                                        img.width = 240
                                        img.height = 180
                                        worksheet.add_image(img, f"I{cell_row}")
                            except Exception as e:
                                st.warning(f"이미지 삽입 실패 (행 {cell_row}): {str(e)}")
                                continue  # 이미지 삽입 실패 시 건너뛰기

                        # URL 열에 하이퍼링크 설정
                        for row_idx in range(2, len(df_youtube) + 2):
                            try:
                                cell = worksheet[f"H{row_idx}"]
                                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                                    cell.hyperlink = cell.value
                                    cell.style = "Hyperlink"
                                    cell.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                            except Exception as e:
                                continue

                        # 특정 열에 줄바꿈 적용
                        for col_letter in ["C", "D", "I"]:  # Title, Channel, Thumbnail
                            for row_idx in range(1, worksheet.max_row + 1):
                                try:
                                    cell = worksheet[f"{col_letter}{row_idx}"]
                                    current_alignment = cell.alignment
                                    new_alignment = Alignment(
                                        horizontal=current_alignment.horizontal if current_alignment else "center",
                                        vertical=current_alignment.vertical if current_alignment else "center",
                                        wrapText=True
                                    )
                                    cell.alignment = new_alignment
                                except Exception as e:
                                    continue

            # 워크북과 관련된 모든 작업이 with 블록 안에서 완료됨
            # (명시적인 save나 close 호출 제거)

            st.download_button(
                label="결과 Excel 파일 다운로드",
                data=excel_buffer.getvalue(),
                file_name="meme_search_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"파일 저장 중 오류 발생: {e}")

    # 결과 표시 (세션 상태 사용)
    if st.session_state.analysis_complete:
        # 각 국가별 결과를 가로로 배치
        for country in st.session_state.results:
            st.header(f"{COUNTRY_FLAGS[country]} {country}")
            
            # 두 개의 컬럼으로 나누기 (비율 조정: 1.2:2.8)
            col1, col2 = st.columns([1.2, 2.8])
            
            # 왼쪽 컬럼: Google Trends 결과
            with col1:
                st.subheader("급상승 검색어")
                trends_df = pd.DataFrame(st.session_state.results[country]["rising_related_queries"])
                if not trends_df.empty:
                    st.dataframe(trends_df, height=600, use_container_width=True)  # 컨테이너 너비 사용
                else:
                    st.write("검색 결과가 없습니다.")
            
            # 오른쪽 컬럼: YouTube Shorts 결과
            with col2:
                st.subheader("YouTube Shorts 결과")
                if country in st.session_state.youtube_results:
                    tabs = st.tabs(list(st.session_state.youtube_results[country].keys()))
                    for tab, (query, shorts) in zip(tabs, st.session_state.youtube_results[country].items()):
                        with tab:
                            if shorts:
                                shorts_df = pd.DataFrame(shorts)
                                st.write(f"**키워드: {query}**")
                                st.dataframe(
                                    shorts_df[['title', 'view_count', 'like_count', 'channel_title', 'url']],
                                    height=500,
                                    use_container_width=True  # 컨테이너 너비 사용
                                )
                            else:
                                st.write("검색 결과가 없습니다.")
            
            # 구분선 제거
            st.write("")  # 간격을 위한 빈 줄 추가

    # 유튜브 영상 다운로드 섹션 추가
    st.divider()
    st.header("📥 YouTube 영상 다운로드")
    
    with st.expander("다운로드 가이드", expanded=True):
        st.markdown("""
        1. 유튜브 url이 포함된 엑셀 파일을 업로드하세요 (`url` 컬럼 필요)
        2. "영상 다운로드 시작" 버튼을 클릭하면 영상들을 다운로드합니다
        3. 다운로드가 완료되면 ZIP 파일로 제공됩니다
        """)
    
    uploaded_video_file = st.file_uploader("유튜브 url이 포함된 엑셀 파일 업로드", type=["xlsx"], key="video_urls")
    
    if uploaded_video_file is not None:
        df_urls = pd.read_excel(uploaded_video_file)
        
        if "url" not in df_urls.columns:
            st.error("엑셀 파일에 'url' 컬럼이 필요합니다.")
        else:
            if st.button("영상 다운로드 시작"):
                st.info("영상을 다운로드 중입니다. 파일 크기에 따라 시간이 걸릴 수 있습니다...")
                
                # 타임스탬프로 임시 폴더 생성
                timestamp = int(time.time())
                download_dir = f"downloaded_videos_{timestamp}"
                os.makedirs(download_dir, exist_ok=True)
                
                # URL 목록 준비
                video_links = []
                for i, row in df_urls.iterrows():
                    url = row["url"]
                    if pd.isna(url) or not isinstance(url, str):
                        continue
                    if "youtube.com/shorts/" in url:
                        video_id = url.split("/")[-1]
                        url = f"https://www.youtube.com/watch?v={video_id}"
                    video_links.append(url)
                
                num_videos = len(video_links)
                st.write(f"총 {num_videos}개의 영상을 다운로드합니다...")
                
                # yt-dlp 옵션 설정
                ydl_opts = {
                    'format': 'bestvideo[ext=mp4][vcodec^=avc]+bestaudio[ext=m4a]/best[ext=mp4]',
                    'outtmpl': os.path.join(download_dir, '%(title)s.%(ext)s'),
                    'merge_output_format': 'mp4',
                    'http_headers': {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'
                    },
                    'force-ipv4': True,
                    'postprocessors': [{
                        'key': 'FFmpegVideoConvertor',
                        'preferedformat': 'mp4'
                    }],
                }
                
                # 다운로드 진행
                failed_list = []
                downloaded_count = 0
                progress_bar = st.progress(0)
                
                for idx, link in enumerate(video_links, start=1):
                    try:
                        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                            ydl.download([link])
                        downloaded_count += 1
                    except Exception as e:
                        failed_list.append({"url": link, "error_msg": str(e)})
                    
                    progress_percent = int(idx / num_videos * 100)
                    progress_bar.progress(progress_percent)
                
                st.success(f"다운로드 완료: 총 {num_videos}개 중 {downloaded_count}개 성공")
                
                # 실패 목록 표시
                if failed_list:
                    st.warning(f"{len(failed_list)}개 영상에서 에러가 발생했습니다.")
                    df_failed = pd.DataFrame(failed_list)
                    with st.expander("다운로드 실패 목록"):
                        st.dataframe(df_failed)
                
                # ZIP 파일 생성
                zip_file_name = f"youtube_videos_{timestamp}.zip"
                with zipfile.ZipFile(zip_file_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root, dirs, files in os.walk(download_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            zipf.write(file_path, arcname=file)
                
                # ZIP 파일 다운로드 버튼
                with open(zip_file_name, "rb") as f:
                    st.download_button(
                        label="ZIP 파일 다운로드",
                        data=f,
                        file_name=zip_file_name,
                        mime="application/zip"
                    )
                
                # 임시 파일/폴더 정리
                try:
                    os.remove(zip_file_name)
                    for f_name in os.listdir(download_dir):
                        os.remove(os.path.join(download_dir, f_name))
                    os.rmdir(download_dir)
                except:
                    pass

if __name__ == "__main__":
    main()
